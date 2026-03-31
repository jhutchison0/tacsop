"""Excel utilities: DataFrame-to-Excel tables, cell indexing, workbook updates."""

import string
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

_ALPHABET = string.ascii_uppercase


def excel_style_index(row: int, col: int) -> str:
    """Convert a 1-based row and column number to an Excel-style cell name.

    Args:
        row: 1-based row number.
        col: 1-based column number.

    Returns:
        Excel cell reference (e.g., "A1", "ZZ100").
    """
    parts: list[str] = []
    c = col
    while c:
        c, rem = divmod(c - 1, 26)
        parts[:0] = _ALPHABET[rem]
    return f"{''.join(parts)}{row}"


def save_excel_table(
    sheets: dict[str, pd.DataFrame],
    file_name: str | Path,
    keep_index: bool = False,
) -> None:
    """Write a dict of DataFrames to Excel, one sheet per key, formatted as tables.

    Args:
        sheets: Mapping of sheet name to DataFrame.
        file_name: Output file path.
        keep_index: If True, include the DataFrame index as the first column.
    """
    writer = pd.ExcelWriter(file_name, engine="xlsxwriter")

    for sheet_name, df in sheets.items():
        df = df.copy(deep=True)
        if keep_index:
            if pd.isna(df.index.name):
                df.index.name = "index"
            df.insert(loc=0, column=df.index.name, value=df.index)

        df.to_excel(
            writer, sheet_name=sheet_name, startrow=1, header=False, index=False
        )

        worksheet = writer.sheets[sheet_name]
        max_row, max_col = df.shape

        column_settings = [{"header": str(h)} for h in df.columns]
        worksheet.add_table(
            0, 0, max_row, max_col - 1, {"columns": column_settings}
        )
        worksheet.set_column(0, max_col - 1, 18)

    writer.close()


def update_excel_workbook(
    file_name: str | Path,
    sheets: dict[str, pd.DataFrame],
    ignore_sheets: list[str] | None = None,
    keep_index: bool = False,
) -> None:
    """Update an existing Excel workbook, replacing sheet data as formatted tables.

    Args:
        file_name: Path to the existing workbook.
        sheets: Mapping of sheet name to replacement DataFrame.
        ignore_sheets: Sheet names to leave untouched.
        keep_index: If True, include the DataFrame index as the first column.
    """
    ignore_sheets = ignore_sheets or []
    wb = load_workbook(file_name)

    for idx_sheet, sheetname in enumerate(wb.sheetnames):
        if sheetname in ignore_sheets:
            continue

        ws = wb[sheetname]
        wb.remove(ws)
        wb.create_sheet(sheetname, idx_sheet)
        ws = wb[sheetname]

        df = sheets[sheetname].copy()
        if keep_index:
            df = df.reset_index(level=0, drop=False)

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        max_row, max_col = df.shape
        idx_final_cell = excel_style_index(max_row + 1, max_col)

        tab = Table(
            displayName=f"Table{idx_sheet + 1}", ref=f"A1:{idx_final_cell}"
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        new_col_length = max(len(str(cell.value)) for cell in ws["A"])
        ws.column_dimensions["A"].width = new_col_length
        ws.add_table(tab)

    wb.save(file_name)
