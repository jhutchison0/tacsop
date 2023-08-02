# -*- coding: utf-8 -*-
"""
Created on Tue Feb 07 11:00:00 2023

@author: jhutchison

"""

# %% Packages
""" Third party and local imports """

import pandas as pd
import string

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

# %% Functions
""" Define functions """


def excel_style_index(row, col):
    """Convert given row and column number to an Excel-style cell name."""
    global alphabet
    list_res = []
    while col:
        col, rem = divmod(col - 1, 26)
        # insert at position 0
        list_res[:0] = alphabet[rem]
    str_res = "".join(list_res)
    idx_excel = f"{str_res}{row}"
    return idx_excel


def save_excel_table(d, file_name, keep_index=False):
    """
    Write a dictionary of dataframes to an Excel Workbook, one dataframe per
    work sheet, saved as an Excel Table.

    Parameters
    ----------
    d : Dictionary, key: sheet_name, value : df
        DESCRIPTION.

    Returns
    -------
    None.

    """
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(file_name, engine="xlsxwriter")

    for sheet_name, df in d.items():
        df = df.copy(deep=True)
        if keep_index:  # == True:
            if pd.isna(df.index.name):  # == True:
                df.index.name = "index"
            df.insert(loc=0, column=df.index.name, value=df.index)

        # Write the dataframe data to XlsxWriter. Turn off default header and
        # index and skip one row to allow us to insert a user defined header.
        df.to_excel(
            writer, sheet_name=sheet_name, startrow=1, header=False, index=False
        )

        # Get the xlsxwriter workbook and worksheet objects.
        # workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Get the dimensions of the dataframe.
        (max_row, max_col) = df.shape

        # Create a list of column headers, to use in add_table().
        column_settings = []
        for header in df.columns:
            column_settings.append({"header": str(header)})

        # Add the table.
        worksheet.add_table(0, 0, max_row, max_col - 1, {"columns": column_settings})

        # Make the columns wider for clarity.
        worksheet.set_column(0, max_col - 1, 18)

    # Close the Pandas Excel writer and output the Excel file.
    # writer.save()
    writer.close()
    return


def update_excel_workbook(
    file_name, ser_data, list_ignore_sheet_names=[], keep_index=False
):
    wb = load_workbook(file_name)

    for idx_sheet, sheetname in enumerate(wb.sheetnames):
        if sheetname in list_ignore_sheet_names:
            continue
        else:
            # Clear worksheet, create a new one
            ws = wb[sheetname]
            wb.remove(ws)

            # create an empty sheet using old index
            wb.create_sheet(sheetname, idx_sheet)
            ws = wb[sheetname]

            if keep_index:
                df = ser_data[sheetname].copy()
            else:
                df = ser_data[sheetname].copy()
                df = df.reset_index(level=0, drop=False)

            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)

            (max_row, max_col) = df.shape
            # zero index
            # shift row by 1
            # col by 0 (added/reset index in code above)
            idx_final_cell = excel_style_index(max_row + 1, max_col)

            tab = Table(displayName=f"Table{idx_sheet + 1}", ref=f"A1:{idx_final_cell}")

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style

            # Adjust first column width:
            new_col_length = max(len(str(cell.value)) for cell in ws["A"])
            # ws.column_dimensions[name].bestFit = True    #I tried this but the result is same
            # Added a extra bit for padding
            ws.column_dimensions["A"].width = new_col_length
            # ws.column_dimensions['A'].width = 40
            ws.add_table(tab)

    wb.save(file_name)
    return


# %% Variables
""" Set script (global) variables """

alphabet = string.ascii_uppercase


# %% Main
""" Test permutations """

if __name__ == "__main__":
    addresses = [
        (1, 1),
        (1, 26),
        (1, 27),
        (1, 52),
        (1, 53),
        (1, 78),
        (1, 79),
        (1, 104),
        (1, 18253),
        (1, 18278),
        (1, 702),  # -> 'ZZ1'
        (1, 703),  # -> 'AAA1'
        (1, 16384),  # -> 'XFD1'
        (1, 35277039),
    ]

    print("({:3}, {:>10}) --> {}".format("row", "col", "Excel"))
    print("==========================")
    for row, col in addresses:
        print("({:3}, {:10,}) --> {!r}".format(row, col, excel_style_index(row, col)))
